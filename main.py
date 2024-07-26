import pathlib, openpyxl
import sys

import openpyxl.worksheet
import openpyxl.worksheet.worksheet

"""Алгоритм такой:
    - Забираем все строки с помощью get_all_rows
    - приводим ячейки к нужному виду с помощью formated_cell
    - создаём хедер с названиями столбцов с помощью create_header
    - собираем все строки в отдельный список с помощью create_result_data
    - заливаем данные в экземпляр xlsx"""

def get_all_rows(sheet:openpyxl.worksheet.worksheet.Worksheet) -> list:
    result = []
    for index, row in enumerate(sheet.iter_rows()):
        result.append([cell.value for cell in row])
        print(f'Считано {index+1} из {sheet.max_row} строк', end='\r')
    else:
        print('\nСтроки считаны', end='\n')
    return result


def formated_cell(row:list) -> list:
    result = row[:]
    if row[2]:
        cell = [x for x in row[2].strip().split('\n') if x and x !=' ']
        if len(cell) > 4:
            cell[-1] = f'{cell[-2]} {cell.pop()}'
        result.extend(cell)
        result.append(row[3])
    return result

def create_result_data(header:list, rows:list):
    print('Собираю итоговые данные')
    result = []
    result.append(header)
    result.extend([formated_cell(row) for row in rows])
    return result


def create_header(sheet:openpyxl.worksheet.worksheet.Worksheet):
    col_names = [x.value for x in list(sheet.rows)[1]]
    col_names.extend(['Контрагент', 'Счет', 'Управление', 'Наименование услуги', 'Цена, руб']) #['Дата', 'Документ', 'Содержание', 'Сумма', 'Контрагент', 'Счет', 'Управление', 'Наименование услуги', 'Цена, руб']
    print(f'Создал названия столбцов {col_names}')
    return col_names

def get_path() -> pathlib.Path:
    name = input('Введите название файла или q для выхода:\n')
    if name == 'q':
        sys.exit(0)
    elif not name:
        print('Имя файла не может быть пустым')
        get_path()
    else:
        path = pathlib.Path().cwd() / name
        if path.exists():
            return path
        else:
            print(f'Файл {path} не существует. Введите имя существующего файла')
            get_path()

def create_new_filepath(file_path:pathlib.Path):
    new_filename = file_path.stem + '-formatted' + file_path.suffix
    print(f'Данные сохранены в файл {new_filename}')
    return new_filename

def main():
    file_path = get_path()
    print(file_path)
    xl = openpyxl.load_workbook(file_path)
    sheet = xl.active
    header = create_header(sheet)
    rows = get_all_rows(sheet)[2:]
    data = openpyxl.Workbook()
    ss = data['Sheet']
    ss.title = 'Счета'
    result_data = create_result_data(header, rows)
    for row in result_data:
        ss.append(row)
    data.save(create_new_filepath(file_path))
    agree = input('Для продолжения нажмите y, n для выхода')
    if agree == 'y':
        pass
    else: sys.exit()

if __name__ == '__main__':
    while True:
        main()